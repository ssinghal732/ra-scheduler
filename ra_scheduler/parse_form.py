"""Read the Google Form export and the roster into AvailabilityData.

This is the module synthetic.py stood in for. Both produce the same object, so
nothing downstream changes when the real one is plugged in.

Two inputs:
  roster xlsx   Email | Name | Tier            (Shivam maintains this)
  form xlsx     the Google Forms response sheet, one row per submission

Decisions this encodes (all Shivam's, 2026-08-24 to 08-26):
  - join on ucsd email first. If that fails, exact full name; if that fails,
    a first name that matches exactly one roster entry. Each fallback FLAGs,
    and if every one fails the run STOPs. (People turned out to have more than
    one ucsd.edu address, so email alone was not the stable key it looked like.)
  - a blackout date is M/D or MM/DD read off the FRONT of each entry, entries
    split on newlines AND commas; anything without a leading date is reported,
    never guessed
  - the two "dates you cannot do" boxes are unioned into one set
  - duplicate submissions: latest timestamp wins, flagged
  - a roster member with no submission STOPS the run
  - all five weekdays marked Class Conflict STOPS the run
  - more than 10 blackout dates is flagged
  - every non-trivial "additional concerns" box is printed for a human to read
  - no NLP anywhere

Column discovery matches a distinctive keyword per field rather than the full
question text (which is a paragraph and will drift), and reports what it
matched so a mismatch is visible in four lines rather than as empty
availability three steps later.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import load_workbook

from .models import (
    FRONT_DESK,
    GAME_ROOM,
    TIER_LRA,
    TIER_NEW,
    TIER_RETURNER,
    RA,
    AvailabilityData,
    Preferences,
    ShiftInstance,
)

WEEKDAYS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
CONFLICT = "class conflict"          # matched case-insensitively as a prefix
MAX_BLACKOUTS_BEFORE_FLAG = 10
TRIVIAL_ANSWERS = {"", "no", "n/a", "na", "none", "nope", "nothing", "-", "nah", "not really"}

# What each field looks for in a header, case-insensitive. Order matters only
# for readability of the report.
COLUMN_KEYS = {
    "timestamp":     ("timestamp",),
    "email":         ("email",),
    "name":          ("ra name",),
    "weekday_dates": ("weekday dates you cannot",),
    "all_dates":     ("all of the dates you cannot",),
    "weekend_pref":  ("weekend shift time preferences",),
    "location":      ("shift location",),
    "concern_wkday": ("concerns related to weekday",),
    "concern_wkend": ("concerns related to weekend",),
}
WEEKDAY_SUFFIX = re.compile(r"\[(Monday|Tuesday|Wednesday|Thursday|Friday)\]\s*$")

STOP, FLAG, READ = "STOP", "FLAG", "READ"
_SEVERITY_ORDER = {STOP: 0, FLAG: 1, READ: 2}

# Every kind of finding, in the order it should print within its severity.
# Most actionable first. Matched against the message text, so the message
# wording and this table live in the same file and change together.
KINDS = [
    (STOP, "has not submitted",                 "on the roster but has not submitted"),
    (STOP, "ALL FIVE weekdays",                 "marked every weekday as a class conflict"),
    (STOP, "resolved to this one roster",       "two submissions matched one roster entry"),
    (STOP, "matches",                           "first name matches more than one roster entry"),
    (STOP, "not on the roster by email",        "not on the roster by email, full name, or first name"),
    (STOP, "duplicate email",                   "roster: duplicate email"),
    (STOP, "malformed email",                   "roster: malformed email"),
    (STOP, "unrecognised tier",                 "roster: unrecognised tier"),
    (FLAG, "the roster has",                    "roster email is stale; matched by name instead"),
    (FLAG, "could not read a date",             "a date the parser could not read"),
    (FLAG, "missing from the all-dates box",    "weekday box lists dates the all-dates box does not"),
    (FLAG, "blackout dates (over",              "more than 10 blackout dates"),
    (FLAG, "weekdays are class conflicts",      "3 or 4 of 5 weekdays are class conflicts"),
    (FLAG, "used more than once",               "same rank used for more than one day"),
    (FLAG, "picked [Open] AND",                 "weekend: [Open] plus a specific day"),
    (FLAG, "outside the quarter",               "blackout dates outside the quarter, ignored"),
    (FLAG, "have no duty anyway",               "blackout dates on days with no duty, harmless"),
    (FLAG, "submitted more than once",          "submitted more than once; later one kept"),
    (FLAG, "neither a rank nor a conflict",     "weekday cell is neither a rank nor a conflict"),
    (FLAG, "outside 1-5",                       "rank outside 1 to 5, ignored"),
    (FLAG, "needed lowercasing",                "roster email needed lowercasing / trimming"),
    (FLAG, "no email",                          "a row with no email was skipped"),
    (READ, "weekday concerns",                  "weekday concerns box"),
    (READ, "weekend concerns",                  "weekend concerns box"),
]


def kind_of(f: "Finding") -> tuple[int, str]:
    """(sort position, label) for a finding, from its message."""
    for i, (sev, needle, label) in enumerate(KINDS):
        if f.severity == sev and needle in f.message:
            return i, label
    return len(KINDS), "other"


@dataclass(frozen=True)
class Finding:
    severity: str
    who: str        # email, or "roster" / "form" for structural findings
    message: str

    def __str__(self) -> str:
        return f"[{self.severity:4s}] {self.who:28s} {self.message}"


@dataclass
class ParseReport:
    matched: dict[str, str] = field(default_factory=dict)   # field -> header it matched
    findings: list[Finding] = field(default_factory=list)
    names: dict[str, str] = field(default_factory=dict)     # email -> display name

    def add(self, severity: str, who: str, message: str) -> None:
        """Findings name people, not emails: Shivam knows names by heart, not
        addresses. `who` may be an email; it is shown as the roster name when
        the roster knows it."""
        self.findings.append(Finding(severity, self.names.get(who, who), message))

    @property
    def must_stop(self) -> bool:
        return any(f.severity == STOP for f in self.findings)

    def summary(self) -> str:
        n = {s: sum(1 for f in self.findings if f.severity == s) for s in (STOP, FLAG, READ)}
        return f"{n[STOP]} stop, {n[FLAG]} flag, {n[READ]} to read"

    def grouped(self) -> list[tuple[str, str, list["Finding"]]]:
        """Findings as (severity, kind label, findings), STOP first, then FLAG,
        then READ; within a severity, most actionable kind first; within a
        kind, by name. This is what the pipeline prints."""
        buckets: dict[tuple[int, int, str], list[Finding]] = {}
        for f in self.findings:
            pos, label = kind_of(f)
            buckets.setdefault((_SEVERITY_ORDER[f.severity], pos, label), []).append(f)
        out = []
        for (sev_i, _, label), fs in sorted(buckets.items()):
            sev = [s for s, i in _SEVERITY_ORDER.items() if i == sev_i][0]
            out.append((sev, label, sorted(fs, key=lambda f: f.who.lower())))
        return out


class ParseError(ValueError):
    """Raised for structural problems that make parsing meaningless."""


# --------------------------------------------------------------------------- #
# small helpers
# --------------------------------------------------------------------------- #
def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def _email(v) -> str:
    return _norm(v).lower()


def _is_trivial(text: str) -> bool:
    return _norm(text).lower().rstrip("!. ") in TRIVIAL_ANSWERS


_DATE_AT_FRONT = re.compile(r"^\s*(\d{1,2})\s*/\s*(\d{1,2})")


def parse_dates(text, year: int) -> tuple[set[date], list[str]]:
    """Dates read off the front of each entry, plus the entries that had none.

    Entries split on newlines and commas: the form asks for one per line, and
    about half the early responses used commas anyway. Everything after the
    date is ignored, so '12/9 (Final)', '12/09(Final)', and '12/9 - Final'
    all read as December 9. 'Oct. 17 (Final)' does not, and is reported.
    """
    found, unreadable = set(), []
    for chunk in re.split(r"[\n,]", str(text or "")):
        chunk = chunk.strip()
        if not chunk or _is_trivial(chunk):
            continue
        m = _DATE_AT_FRONT.match(chunk)
        if not m:
            unreadable.append(chunk)
            continue
        try:
            found.add(date(year, int(m.group(1)), int(m.group(2))))
        except ValueError:
            unreadable.append(chunk)
    return found, unreadable


def _rank(v) -> int | None:
    """A 1-5 rank from a cell that may hold a float, an int, or a string."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = _norm(v)
    return int(s) if s.isdigit() else None


def _is_conflict(v) -> bool:
    return _norm(v).lower().startswith(CONFLICT)


# --------------------------------------------------------------------------- #
# roster
# --------------------------------------------------------------------------- #
_TIER_WORDS = {
    "lra": TIER_LRA, "lead": TIER_LRA,
    "returner": TIER_RETURNER, "returning": TIER_RETURNER, "return": TIER_RETURNER,
    "new": TIER_NEW, "first": TIER_NEW,
}


def load_roster(path: str, report: ParseReport) -> list[RA]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True) if any(_norm(c) for c in r)]
    if not rows:
        raise ParseError(f"roster {path}: no rows")

    hdr = [_norm(c).lower() for c in rows[0]]
    col = {}
    for key in ("email", "name", "tier"):
        hits = [i for i, h in enumerate(hdr) if key in h]
        if not hits:
            raise ParseError(f"roster: no column containing '{key}' in {hdr}")
        col[key] = hits[0]
    report.matched["roster"] = f"{path} ({len(rows) - 1} rows)"

    roster, seen = [], {}
    for r in rows[1:]:
        email, name, tier_raw = _email(r[col["email"]]), _norm(r[col["name"]]), _norm(r[col["tier"]])
        if not email:
            report.add(FLAG, "roster", f"row with no email skipped (name {name!r})")
            continue
        if email in seen:
            report.add(STOP, "roster", f"duplicate email {email}")
            continue
        if "@" not in email:
            report.add(STOP, "roster", f"malformed email {email!r}")
            continue
        report.names[email] = name or email
        if _norm(r[col["email"]]) != email:
            report.add(FLAG, email, "roster email needed lowercasing / trimming to match")
        tier = _TIER_WORDS.get(tier_raw.lower().split()[0] if tier_raw else "")
        if tier is None:
            report.add(STOP, email, f"unrecognised tier {tier_raw!r} (want LRA / Returner / New)")
            continue
        seen[email] = True
        roster.append(RA(ra_id=email, name=name or email, tier=tier))

    counts = {t: sum(1 for ra in roster if ra.tier == t) for t in (TIER_LRA, TIER_RETURNER, TIER_NEW)}
    report.matched["roster tiers"] = (f"{counts[TIER_LRA]} LRA / {counts[TIER_RETURNER]} returner / "
                                      f"{counts[TIER_NEW]} new = {len(roster)}")
    return roster


# --------------------------------------------------------------------------- #
# form
# --------------------------------------------------------------------------- #
def _find_columns(hdr: list, report: ParseReport) -> dict:
    low = [_norm(h).lower() for h in hdr]
    col: dict = {}
    for key, needles in COLUMN_KEYS.items():
        hits = [i for i, h in enumerate(low) if any(n in h for n in needles)]
        if not hits:
            raise ParseError(f"form: no column for '{key}' (looked for {needles})")
        col[key] = hits[0]
        report.matched[key] = _norm(hdr[hits[0]])[:60]

    col["rank"] = {}
    for i, h in enumerate(hdr):
        m = WEEKDAY_SUFFIX.search(_norm(h))
        if m:
            col["rank"][m.group(1)] = i
    missing = [d for d in WEEKDAYS if d not in col["rank"]]
    if missing:
        raise ParseError(f"form: no ranking column for {missing}; want headers ending [Monday]..[Friday]")
    report.matched["weekday ranks"] = "5 columns ending [Monday] .. [Friday]"
    return col


def _parse_weekend(text: str) -> tuple[set[str], set[str]]:
    """'[Sundays] ..., [Morning] ..., [Evening] ...' -> ({'Sunday'}, {'Morning','Evening'})."""
    labels = set(re.findall(r"\[([A-Za-z]+)\]", str(text or "")))
    days = {d[:-1] if d.endswith("s") else d for d in labels if d in ("Saturdays", "Sundays")}
    times = {t for t in labels if t in ("Morning", "Afternoon", "Evening")}
    if "Open" in labels:
        days = set()
    if len(times) == 3:
        times = set()          # all three is no preference
    return days, times


def _parse_location(text: str) -> str:
    s = _norm(text).lower()
    if "either" in s or not s:
        return ""
    if "front desk" in s:
        return FRONT_DESK
    if "game room" in s:
        return GAME_ROOM
    return ""


def load_form(
    path: str,
    shifts: list[ShiftInstance],
    roster: list[RA],
    report: ParseReport,
    sheet: str | None = None,
) -> AvailabilityData:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ParseError(f"form {path}: header only, no responses")
    col = _find_columns(list(rows[0]), report)
    report.matched["form"] = f"{path} sheet {ws.title!r} ({len(rows) - 1} responses)"

    year = shifts[0].date.year
    first, last = min(s.date for s in shifts), max(s.date for s in shifts)
    duty_dates = {s.date for s in shifts}
    by_email = {ra.ra_id: ra for ra in roster}
    by_name = {_norm(ra.name).lower(): ra for ra in roster}
    by_first: dict[str, list[RA]] = {}
    for ra in roster:
        by_first.setdefault(_norm(ra.name).lower().split()[0], []).append(ra)

    def resolve(email: str, name: str) -> RA | None:
        """Email, then full name, then unique first name. Flag every fallback."""
        if email in by_email:
            return by_email[email]
        who = f"{_norm(name) or '(no name)'} <{email}>"     # not on the roster: show both
        full = _norm(name).lower()
        if full in by_name:
            ra = by_name[full]
            report.add(FLAG, who, f"email not on the roster; matched by full name. "
                                  f"the roster has {ra.ra_id} for them, update it")
            return ra
        fn = full.split()[0] if full else ""
        hits = by_first.get(fn, [])
        if len(hits) == 1:
            report.add(FLAG, who, f"email not on the roster; matched by first name only to "
                                  f"{hits[0].name}. the roster has {hits[0].ra_id} for them, update it")
            return hits[0]
        if len(hits) > 1:
            report.add(STOP, who, f"first name {fn!r} matches {len(hits)} roster entries; "
                                  f"cannot tell which. Fix the roster email.")
        else:
            report.add(STOP, who, "not on the roster by email, full name, or first name. "
                                  "Add them or fix the roster.")
        return None

    # --- one row per person: latest timestamp wins ---
    latest: dict[str, tuple] = {}
    for r in rows[1:]:
        email = _email(r[col["email"]])
        if not email:
            report.add(FLAG, "form", "response with no email skipped")
            continue
        ts = r[col["timestamp"]]
        ts = ts if isinstance(ts, datetime) else datetime.min
        if email in latest:
            keep_new = ts >= latest[email][0]
            report.add(FLAG, email, f"submitted more than once; keeping the "
                                    f"{'later' if keep_new else 'earlier'} one")
            if not keep_new:
                continue
        latest[email] = (ts, r)

    available: dict[str, set[str]] = {}
    blackouts: dict[str, set[date]] = {}
    prefs: dict[str, Preferences] = {}

    for form_email, (_, r) in latest.items():
        ra = resolve(form_email, r[col["name"]])
        if ra is None:
            continue
        email = ra.ra_id           # everything downstream keys on the ROSTER id
        if email in available:
            report.add(STOP, email, "two different form submissions resolved to this one roster entry")
            continue

        # weekday ranking: rank or conflict per day
        conflicts, ranks = set(), {}
        for dow in WEEKDAYS:
            v = r[col["rank"][dow]]
            if _is_conflict(v):
                conflicts.add(dow)
            else:
                k = _rank(v)
                if k is None:
                    report.add(FLAG, email, f"{dow}: neither a rank nor a conflict ({v!r})")
                elif not 1 <= k <= 5:
                    report.add(FLAG, email, f"{dow}: rank {k} is outside 1-5, ignored")
                else:
                    ranks[dow] = k
        if len(conflicts) == 5:
            report.add(STOP, email, "marked ALL FIVE weekdays as class conflicts; weekend-only. Check with them.")
        elif len(conflicts) >= 3:
            report.add(FLAG, email, f"{len(conflicts)} of 5 weekdays are class conflicts")
        dupes = {k for k in ranks.values() if list(ranks.values()).count(k) > 1}
        if dupes:
            report.add(FLAG, email, f"rank(s) {sorted(dupes)} used more than once")

        # blackout dates: union of both boxes
        wk_dates, bad1 = parse_dates(r[col["weekday_dates"]], year)
        all_dates, bad2 = parse_dates(r[col["all_dates"]], year)
        for frag in bad1 + bad2:
            report.add(FLAG, email, f"could not read a date from {frag[:40]!r}")
        only_in_wk = wk_dates - all_dates
        if only_in_wk:
            report.add(FLAG, email, f"{len(only_in_wk)} date(s) in the weekday box are missing "
                                    f"from the all-dates box; unioned")
        black = wk_dates | all_dates
        outside = {d for d in black if not first <= d <= last}
        if outside:
            report.add(FLAG, email, f"{len(outside)} blackout date(s) fall outside the quarter, e.g. "
                                    f"{min(outside):%m/%d}; ignored")
            black -= outside
        idle = {d for d in black if d not in duty_dates}
        if idle:
            report.add(FLAG, email, f"{len(idle)} blackout date(s) have no duty anyway, e.g. {min(idle):%m/%d}")
        if len(black) > MAX_BLACKOUTS_BEFORE_FLAG:
            report.add(FLAG, email, f"{len(black)} blackout dates (over {MAX_BLACKOUTS_BEFORE_FLAG})")

        # weekend + location preferences
        wk_days, wk_times = _parse_weekend(r[col["weekend_pref"]])
        raw = str(r[col["weekend_pref"]] or "")
        if "[Open]" in raw and ("[Saturdays]" in raw or "[Sundays]" in raw):
            report.add(FLAG, email, "weekend: picked [Open] AND a specific day; treating as open")
        location = _parse_location(r[col["location"]])

        # concerns: a human reads these, always
        for key, label in (("concern_wkday", "weekday concerns"), ("concern_wkend", "weekend concerns")):
            text = _norm(r[col[key]])
            if not _is_trivial(text):
                report.add(READ, email, f"{label}: {text}")

        # availability by subtraction
        ok = {s.key for s in shifts
              if s.date not in black
              and not (s.shift == "Evening (Weekday)" and s.dow in conflicts)}
        available[email] = ok
        blackouts[email] = black
        prefs[email] = Preferences(weekday_rank=ranks, weekend_days=wk_days,
                                   weekend_times=wk_times, location=location)

    # --- everyone on the roster must have answered ---
    missing = [ra for ra in roster if ra.ra_id not in available]
    for ra in missing:
        report.add(STOP, ra.ra_id, f"on the roster ({ra.tier}) but has not submitted the form")
    report.matched["responses used"] = f"{len(available)} of {len(roster)} roster members"

    return AvailabilityData(roster=roster, available=available,
                            blackout_dates=blackouts, preferences=prefs)


def load(roster_path: str, form_path: str, shifts: list[ShiftInstance],
         sheet: str | None = None) -> tuple[AvailabilityData, ParseReport]:
    """The one call run_pipeline makes."""
    report = ParseReport()
    roster = load_roster(roster_path, report)
    data = load_form(form_path, shifts, roster, report, sheet=sheet)
    return data, report
