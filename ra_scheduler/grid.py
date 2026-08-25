"""Read the duty-schedule grid (xlsx) into ShiftInstances.

The grid is the source of truth for WHICH (date, shift) need coverage and how
many seats each has. Holiday rows are returned separately so the exporter can
reproduce them; they need no coverage.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from .models import STAFFING, ShiftInstance

_EXCEL_EPOCH = datetime(1899, 12, 30)


@dataclass(frozen=True)
class HolidayRow:
    date: date
    dow: str
    shift: str
    time: str
    rounds: str
    week: str


def _to_date(v) -> date | None:
    if hasattr(v, "date"):  # datetime
        return v.date()
    if isinstance(v, (int, float)):  # Excel serial
        return (_EXCEL_EPOCH + timedelta(days=int(v))).date()
    return None


def load_grid(path: str, sheet: str = "Fall Duty Schedule") -> tuple[list[ShiftInstance], list[HolidayRow]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))

    hdr_idx = next(i for i, r in enumerate(rows) if r and "Date" in r and "Shift" in r)
    col = {name: i for i, name in enumerate(rows[hdr_idx]) if isinstance(name, str)}

    def get(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    shifts: list[ShiftInstance] = []
    holidays: list[HolidayRow] = []
    week = ""  # forward-fill: the grid writes the week label on its first row only
    for row in rows[hdr_idx + 1 :]:
        d = _to_date(get(row, "Date"))
        shift = get(row, "Shift")
        if d is None or shift not in STAFFING:
            continue  # roster side-table rows, totals, blanks
        w = get(row, "Week")
        if w is not None and str(w).strip() != "":
            week = str(int(w)) if isinstance(w, float) else str(w)
        common = dict(
            date=d,
            dow=str(get(row, "Day of the Week")).strip(),
            shift=shift,
            time=str(get(row, "Time")),
            rounds=str(get(row, "Duty Round Location") or "").strip(),
            week=week,
        )
        fd_primary = get(row, "Front Desk Primary")
        if isinstance(fd_primary, str) and fd_primary.strip().lower() == "holiday":
            holidays.append(HolidayRow(**common))
        else:
            shifts.append(ShiftInstance(sid=len(shifts), **common))
    return shifts, holidays
