"""Write the filled schedule to xlsx in the team's exact column shape.

Matches the grid the duty leads already use: same headers, holiday rows kept in
place, week label written once per week, Exceptions listing date-blackout RAs.
Values only (no formulas), Arial throughout.
"""
from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import RA, AvailabilityData, ShiftInstance, WEEKEND_TIME_LABEL
from .grid import HolidayRow
from .roles import FD_P, FD_S, GR_P, GR_S

HEADERS = (
    "Week", "Date", "Day of the Week", "Shift", "Time",
    "Front Desk Primary", "Front Desk Secondary",
    "Game Room Primary", "Game Room Secondary",
    "Duty Round Location", "Exceptions",
)
_SHIFT_ORDER = {"Evening (Weekday)": 0, "Morning": 0, "Afternoon": 1, "Evening (Weekend)": 2}


SYNTHETIC_TAG = "SYNTHETIC"


def synthetic_path(path: str) -> str:
    """Force SYNTHETIC into a filename so an invented schedule cannot be mailed out.

    A run on synthetic availability produces a file that is correct in every
    visible way and staffed by people who do not exist. The filename is the
    only thing standing between that file and someone's inbox, so the caller
    does not get to choose whether it says so. Idempotent: a name that already
    carries the tag is returned unchanged.
    """
    p = Path(path)
    if SYNTHETIC_TAG.lower() in p.stem.lower():
        return path
    return str(p.with_name(f"{p.stem}_{SYNTHETIC_TAG}{p.suffix}"))


def _exceptions_by_date(data: AvailabilityData) -> dict[date, str]:
    name_of = {ra.ra_id: ra.name for ra in data.roster}
    out: dict[date, list[str]] = {}
    for rid, days in data.blackout_dates.items():
        for d in days:
            out.setdefault(d, []).append(name_of[rid])
    return {d: ", ".join(sorted(names)) for d, names in out.items()}


AVAIL_HEADERS = (
    "RA", "Tier", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
    "Dates they cannot work", "Weekend pref", "Desk pref",
    "Shifts they could work", "Shifts assigned", "Weekday / Weekend",
    "Weekday evenings they got", "On a 1st/2nd choice day",
)
WEEKDAYS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
_UNAVAILABLE = "Class Conflict/Unavailable"


def _availability_rows(shifts, data: AvailabilityData, roles) -> list[tuple]:
    """One row per RA, in the shape the availability form collects.

    The weekday columns are DERIVED, not stored: an RA counts as unavailable on
    a weekday when there is no non-blacked-out weekday evening on that day they
    can work. That is the same signal the form's ranking matrix carries, read
    back out of the availability set so this sheet cannot drift from what the
    solver actually saw.
    """
    assigned: dict[str, int] = {}
    for table in roles.values():
        for rid in table.values():
            assigned[rid] = assigned.get(rid, 0) + 1

    # weekday evenings each RA got, and how many landed on a 1st/2nd choice day
    kept: dict[str, list[int]] = {}
    worked: dict[str, Counter] = {}
    by_sid = {s.sid: s for s in shifts}
    for sid, table in roles.items():
        s = by_sid.get(sid)
        if not s or s.shift != "Evening (Weekday)":
            continue
        for rid in table.values():
            rank = data.prefs(rid).weekday_rank.get(s.dow)
            got, tot = kept.setdefault(rid, [0, 0])
            kept[rid] = [got + (1 if rank and rank <= 2 else 0), tot + 1]
            worked.setdefault(rid, Counter())[s.dow] += 1

    out = []
    for ra in data.roster:
        free = data.available.get(ra.ra_id, set())
        blackout = data.blackout_dates.get(ra.ra_id, set())
        prefs = data.prefs(ra.ra_id)
        cells = []
        for dow in WEEKDAYS:
            candidates = [s for s in shifts
                          if s.dow == dow and s.shift == "Evening (Weekday)"
                          and s.date not in blackout]
            if not candidates:
                cells.append("")                      # no such shift this quarter
            elif any(s.key in free for s in candidates):
                rank = prefs.weekday_rank.get(dow)
                cells.append(f"{rank}" if rank else "Available")
            else:
                cells.append(_UNAVAILABLE)
        wk = " / ".join(x for x in (
            ", ".join(sorted(prefs.weekend_days)) or "any day",
            ", ".join(sorted(prefs.weekend_times)) or "any time") if x)
        got, tot = kept.get(ra.ra_id, [0, 0])
        days = worked.get(ra.ra_id, Counter())
        # "Mon x4, Wed x1", most-worked day first: their duty day at a glance
        day_text = ", ".join(f"{d[:3]} x{n}" for d, n in
                             sorted(days.items(), key=lambda kv: (-kv[1], kv[0])))
        out.append((
            ra.name,
            ra.tier,   # LRA / returner / new: this is what sets their target
            *cells,
            ", ".join(d.strftime("%m/%d") for d in sorted(blackout)),
            wk,
            prefs.location or "either",
            len(free),
            assigned.get(ra.ra_id, 0),
            f"{sum(days.values())} / {assigned.get(ra.ra_id, 0) - sum(days.values())}",
            day_text or "none",
            f"{got}/{tot}" if tot else "-",
        ))
    return out


def write_schedule(
    path: str,
    shifts: list[ShiftInstance],
    holidays: list[HolidayRow],
    roles: dict[int, dict[str, str]],
    data: AvailabilityData,
    sheet_name: str = "Fall Duty Schedule",
) -> None:
    name_of = {ra.ra_id: ra.name for ra in data.roster}
    exceptions = _exceptions_by_date(data)

    # interleave duty rows and holiday rows in calendar order
    rows: list[tuple] = []
    for s in shifts:
        t = roles[s.sid]
        rows.append((s.date, _SHIFT_ORDER[s.shift], s.week, s.dow, s.shift, s.time,
                     name_of[t[FD_P]], name_of[t[FD_S]],
                     name_of.get(t.get(GR_P, ""), ""), name_of.get(t.get(GR_S, ""), ""),
                     s.rounds))
    for h in holidays:
        rows.append((h.date, _SHIFT_ORDER[h.shift], h.week, h.dow, h.shift, h.time,
                     "Holiday", "Holiday",
                     "Holiday" if h.shift.startswith("Evening") else "",
                     "Holiday" if h.shift.startswith("Evening") else "",
                     h.rounds))
    rows.sort(key=lambda r: (r[0], r[1]))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    body_font = Font(name="Arial")
    holiday_fill = PatternFill("solid", fgColor="FCE4D6")

    ws.append(HEADERS)
    for c in ws[1]:
        c.font, c.fill = header_font, header_fill
        c.alignment = Alignment(horizontal="center")

    prev_week, prev_date = None, None
    for (d, _, week, dow, shift, time, fdp, fds, grp, grs, rounds) in rows:
        ws.append((
            week if week != prev_week else "",
            d, dow, shift, time, fdp, fds, grp, grs, rounds,
            exceptions.get(d, "") if d != prev_date else "",  # once per date
        ))
        for c in ws[ws.max_row]:
            c.font = body_font
        ws.cell(ws.max_row, 2).number_format = "MM/DD/YYYY"
        if fdp == "Holiday":
            for c in ws[ws.max_row]:
                c.fill = holiday_fill
        prev_week, prev_date = week, d

    for col_cells, width in zip(ws.columns, (6, 12, 14, 18, 11, 17, 17, 17, 17, 26, 40)):
        ws.column_dimensions[col_cells[0].column_letter].width = width
    ws.freeze_panes = "A2"

    # Second tab: what the solver was given. Lets anyone reading the schedule
    # check an assignment against the availability behind it without a rerun.
    aw = wb.create_sheet("Availability")
    aw.append(AVAIL_HEADERS)
    for c in aw[1]:
        c.font, c.fill = header_font, header_fill
        c.alignment = Alignment(horizontal="center")
    for row in _availability_rows(shifts, data, roles):
        aw.append(row)
        for c in aw[aw.max_row]:
            c.font = body_font
        for j in (11, 12, 13, 15):
            aw.cell(aw.max_row, j).alignment = Alignment(horizontal="center")
        for j in range(3, 8):
            if aw.cell(aw.max_row, j).value == _UNAVAILABLE:
                aw.cell(aw.max_row, j).fill = holiday_fill
    for col_cells, width in zip(aw.columns,
                                (12, 11, 22, 22, 22, 22, 22, 32, 26, 12, 20, 16, 18, 30, 22)):
        aw.column_dimensions[col_cells[0].column_letter].width = width
    aw.freeze_panes = "C2"

    wb.save(path)
